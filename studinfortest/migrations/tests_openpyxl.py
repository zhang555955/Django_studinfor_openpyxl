#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @File  : tests_openpyxl.py
# @Author: 橘子
# @Date  : 2020/12/29
# @Desc  : execise

import openpyxl

def write_excel():

    #创建工作簿
    f = openpyxl.Workbook()

    #创建表
    sheet1 = f.create_sheet('业务统计')

    row0 = [u'业务', u'状态', u'北京', u'上海', u'广州', u'深圳', u'状态小计', u'合计']
    column0 = [u'机票', u'船票', u'火车标', u'汽车票', u'其它']
    status = [u'预定', u'出票', u'退票', u'业务小计']

    #生成第一行
    for i in range(len(row0)):
        j = i + 1
        sheet1.cell(row=1, column=j).value = row0[i]

   #生成第一列和最后一列（合并4行）
    i, j = 1, 1
    while i < 4 * len(column0) and j < len(column0):
        for j in range(len(column0)):
            #填写第一列数据
            sheet1.cell(row=i+1, column=1).value = column0[j]
            #第一列合并单元格
            sheet1.merge_cells(None, start_row=i+1, start_column=1, end_row=i+4, end_column=1)
            #最后一列“合计”合并单元格，4个合一个
            sheet1.merge_cells(None, start_row=i+1, start_column=8, end_row=i+4, end_column=8)
            i += 4
            j += 1

    #最后一行22行的数据写入"合计"
    sheet1.cell(row=22, column=1).value = u'合计'
    #最后一行22行的第1列和第二列合并
    sheet1.merge_cells(None, start_row=22, start_column=1, end_row=22, end_column=2)

    #生成第二列
    i = 1
    while i < 4*len(column0):
        for j in range(len(status)):
            sheet1.cell(row=j+i+1, column=2).value = status[j]
        i += 4
        f.save('demo1.xls')

    '''
        创建第二个sheet:sheet2
    '''
    #创建sheet2
    sheet2 = f.create_sheet('资料')

    row0 = [u'姓名',u'年龄',u'出生日期',u'爱好',u'关系']
    column0 = [u'小杰',u'小胖',u'小明',u'大神',u'大仙',u'小敏',u'无名']

    #生成第一行
    for i in range(len(row0)):
        sheet2.cell(row=1, column=i+1).value = row0[i]

    #生成第一例
    for i in range(len(column0)):
        sheet2.cell(row=i+2, column=1).value = column0[i]

    sheet2.cell(row=2, column=3).value = '1991/11/11'
    sheet2.cell(row=8, column=3).value = '暂无'
    sheet2.merge_cells(None, start_row=8, start_column=3, end_row=8, end_column=5)

    sheet2.cell(row=2, column=5).value = '好朋友'
    sheet2.merge_cells(None, start_row=2, start_column=5, end_row=3, end_column=5)

    f.save('demo1.xls')

if __name__ =='__main__':
    write_excel()
    print('写入成功')

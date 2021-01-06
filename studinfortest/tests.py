# -*- coding: utf-8 -*-
import string

from django.test import TestCase

# Create your tests here.
#  四、如果开启了鉴权，那么如何获取get接口
# 在开头里增加语句：from requests.auth import HTTPBasicAuth
# 另外，修改此句:即增加了auth的用户名与密码
#
# self.r = requests.get('http://localhost:8080/jenkins/api/json?tree=jobs[name]'，auth=('admin','admin')

import requests
import openpyxl
from js_asset import js
import json


def postRequest(url,data):
    responses = requests.post(url, data)
    a_result = responses.json()
    if a_result:
      if a_result['success'] != '0':
        # value1 = json.dumps(a_result['result'], indent=2, ensure_ascii=False)
        value1 = a_result['result']
        return (value1)

      else:
        return (a_result['msgid']+' '+a_result['msg'])
    else:
      return ('Request nowapi fail.')


def write_excel_xlsx():
    #创建工作簿
    f = openpyxl.Workbook()
    #创建表
    sheet = f.create_sheet('天气接口格式测试表')
    # sheet.cell(row=1, column=1).value = 1
    sheet.sheet_properties.tabColor = "1072BA"

    value1 = postRequest(url, data)
    for k in range(len(value1)):
        v = value1[k]
        v1 = format(v)
        print(v1)
        print(type(v1))
        # value2 = format(value1)
        sheet.cell(row=k+1, column=1).value = k
        sheet.cell(row=k+1, column=2).value = v1
        f.save('天气接口测试工作簿.xlsx')

    print("xlsx格式表格写入数据成功！")


if __name__ == '__main__':
    url = 'http://api.k780.com'
    data = {
        'app': 'weather.future',
        'weaid': '北京',
        'ag': 'today',          #,'futureDay','lifeIndex','futureHour',
        'appkey': '56366',         #APPKEY
        'sign': 'a9682f8c0dc8e8f32aa5597f4649c700',
        'format': 'json',
    }

    write_excel_xlsx()


    # values = json.dumps(value1, indent=2, ensure_ascii=False)
    # value3 = value2.split(':')



    # print(value3)
    # print(len(value3))
    # print(type(value3))

    # for i in range(len(value2)):
    #     j = i + 1
    #     sheet.cell(row=1, column=j).value = value2[i]




